"""
Convert the Excel file to CSV format
This handles the corrupted Excel file structure
"""

import pandas as pd
from io import StringIO

# The data content from the Excel file
data_content = """Gallstone Status	Age	Gender	Comorbidity	Coronary Artery Disease (CAD)	Hypothyroidism	Hyperlipidemia	Diabetes Mellitus (DM)	Height	Weight	Body Mass Index (BMI)	Total Body Water (TBW)	Extracellular Water (ECW)	Intracellular Water (ICW)	Extracellular Fluid/Total Body Water (ECF/TBW)	Total Body Fat Ratio (TBFR) (%)	Lean Mass (LM) (%)	Body Protein Content (Protein) (%)	Visceral Fat Rating (VFR)	Bone Mass (BM)	Muscle Mass (MM)	Obesity (%)	Total Fat Content (TFC)	Visceral Fat Area (VFA)	Visceral Muscle Area (VMA) (Kg)	Hepatic Fat Accumulation (HFA)	Glucose	Total Cholesterol (TC)	Low Density Lipoprotein (LDL)	High Density Lipoprotein (HDL)	Triglyceride	Aspartat Aminotransferaz (AST)	Alanin Aminotransferaz (ALT)	Alkaline Phosphatase (ALP)	Creatinine	Glomerular Filtration Rate (GFR)	C-Reactive Protein (CRP)	Hemoglobin (HGB)	Vitamin D
0	50	0	0	0	0	0	0	185	92.8	27.1	52.9	21.2	31.7	40	19.2	80.84	18.88	9	3.7	71.4	23.4	17.8	10.6	39.7	0	102	250	175	40	134	20	22	87	0.82	112.47	0	16	33"""

# Read the tab-separated data
df = pd.read_csv(StringIO(data_content), sep='\t', nrows=1)

# Now read the full Excel file using a different approach
try:
    # Try reading with xlrd engine for older Excel files
    df_full = pd.read_excel('../data/dataset-uci.xlsx', engine='openpyxl', sheet_name=None)
    print("Available sheets:", list(df_full.keys()))
    
    # Get the first available sheet
    if df_full:
        first_sheet = list(df_full.keys())[0]
        df = df_full[first_sheet]
        print(f"Loaded sheet: {first_sheet}")
        print(f"Shape: {df.shape}")
except Exception as e:
    print(f"Error with openpyxl: {e}")
    print("\nTrying alternative method...")
    
    # Alternative: Use openpyxl directly
    from openpyxl import load_workbook
    
    wb = load_workbook('../data/dataset-uci.xlsx', data_only=True)
    print(f"Workbook sheets: {wb.sheetnames}")
    
    if wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
        
        # Extract data from worksheet
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data[1:], columns=data[0])
        print(f"Extracted data shape: {df.shape}")
    else:
        print("No sheets found in workbook")
        df = None

# Save to CSV if successful
if df is not None and not df.empty:
    output_path = '../data/dataset-uci.csv'
    df.to_csv(output_path, index=False)
    print(f"\n[SUCCESS] Converted to CSV: {output_path}")
    print(f"Final shape: {df.shape[0]} rows x {df.shape[1]} columns")
    print(f"\nFirst few column names:")
    for i, col in enumerate(df.columns[:10], 1):
        print(f"  {i}. {col}")
else:
    print("[ERROR] Failed to extract data")

