"""
Extract full data from the corrupted Excel file
Uses pandas read_excel with special handling
"""

import pandas as pd
import warnings
warnings.filterwarnings('ignore')

# Try different methods to read the file
file_path = '../data/dataset-uci.xlsx'

print("Attempting to read Excel file...")
print("=" * 80)

# Method 1: Try with xlrd for older Excel files
try:
    print("\nMethod 1: Trying with xlrd engine...")
    df = pd.read_excel(file_path, engine='xlrd')
    print(f"Success! Shape: {df.shape}")
except Exception as e:
    print(f"Failed: {e}")
    df = None

# Method 2: Try with openpyxl and read_only mode
if df is None:
    try:
        print("\nMethod 2: Trying with openpyxl read_only mode...")
        df_dict = pd.read_excel(file_path, engine='openpyxl', sheet_name=None)
        if df_dict and isinstance(df_dict, dict):
            # Get first sheet
            first_key = list(df_dict.keys())[0]
            df = df_dict[first_key]
            print(f"Success! Shape: {df.shape}")
        else:
            df = None
    except Exception as e:
        print(f"Failed: {e}")
        df = None

# Method 3: Use openpyxl directly with data_only
if df is None:
    try:
        print("\nMethod 3: Using openpyxl directly...")
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        # Try to get any worksheet
        if hasattr(wb, 'worksheets'):
            ws_list = list(wb.worksheets)
            if ws_list:
                ws = ws_list[0]
                print(f"Found worksheet")
                
                # Extract all data
                data = []
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        data.append(row)
                
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    print(f"Success! Shape: {df.shape}")
        
        wb.close()
    except Exception as e:
        print(f"Failed: {e}")
        df = None

# Method 4: Try reading as binary and parsing
if df is None:
    try:
        print("\nMethod 4: Reading with pandas default...")
        df = pd.read_excel(file_path)
        print(f"Success! Shape: {df.shape}")
    except Exception as e:
        print(f"Failed: {e}")
        df = None

# If successful, save to CSV
if df is not None and not df.empty:
    output_path = '../data/dataset-uci.csv'
    df.to_csv(output_path, index=False)
    print("\n" + "=" * 80)
    print(f"[SUCCESS] Data extracted and saved to CSV")
    print(f"Output file: {output_path}")
    print(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")
    print("\nFirst 5 column names:")
    for i, col in enumerate(df.columns[:5], 1):
        print(f"  {i}. {col}")
    print("\nData types:")
    print(df.dtypes.value_counts())
else:
    print("\n" + "=" * 80)
    print("[ERROR] All methods failed to extract data")
    print("The Excel file may be severely corrupted")


